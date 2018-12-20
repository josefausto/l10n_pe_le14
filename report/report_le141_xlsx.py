from odoo import models
from io import BytesIO
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# from openpyxl.writer.excel import save_virtual_workbook

class PartnerXlsx(models.AbstractModel):
    _name = 'report.l10n_pe_le14.export_le141_xlsx'
    _inherit = 'report.report_xlsx.abstract'

    def create_xlsx_report(self, docids, data):
        objs = self.env[self.env.context.get('active_model')].browse(docids)
        file_data = BytesIO()
        workbook = load_workbook("%s/%s" %(os.path.dirname(__file__).replace('\\','/'),"formato141.xlsx"))
        workbook.template = False
        self.generate_xlsx_report(workbook, data, objs)
        workbook.save(file_data)
        file_data.seek(0)
        return file_data.read(), 'xlsx'

    def generate_xlsx_report(self, workbook, data, objs):
        worksheet = workbook.active
        worksheet["B3"] = objs.start_date and datetime.strptime(objs.start_date, "%Y-%m-%d").strftime('%Y%m') or ''
        worksheet["B4"] = objs.company_id.partner_id.doc_number or ""
        worksheet["E5"] = objs.company_id.name or ""
        row = 12
        for line_id in objs.line_ids:
            worksheet["A%d"%row] = line_id.sequence or ""
            worksheet["B%d"%row] = line_id.date and datetime.strptime(line_id.date, "%Y-%m-%d").date() or ''
            worksheet["B%d"%row].number_format = "dd-mm-yyyy"
            worksheet["C%d"%row] = line_id.due_date and datetime.strptime(line_id.due_date, "%Y-%m-%d").date() or ''
            worksheet["C%d"%row].number_format = "dd-mm-yyyy"
            worksheet["D%d"%row] = line_id.document_type or "-"
            
            worksheet["E%d"%row] = line_id.series or ""
            worksheet["F%d"%row] = line_id.name or ""
            
            worksheet["G%d"%row] = line_id.doc_type or ""
            worksheet["H%d"%row] = line_id.doc_number or ""
            worksheet["I%d"%row] = line_id.partner_name or ""
            
            worksheet["J%d"%row] = line_id.amount_export or ""
            worksheet["K%d"%row] = line_id.amount or ""
            worksheet["L%d"%row] = line_id.amount_exonerated or ""
            
            worksheet["M%d"%row] = line_id.amount_unaffected or ""
            worksheet["N%d"%row] = line_id.amount_isc or ""
            
            worksheet["O%d"%row] = line_id.amount_igv or ""
            worksheet["P%d"%row] = line_id.amount_other or ""
            
            worksheet["Q%d"%row] = line_id.amount_total or ""
            
            worksheet["R%d"%row] = line_id.rate or ""
            
            worksheet["S%d"%row] = line_id.related_date  and datetime.strptime(line_id.related_date, "%Y-%m-%d").date() or ''
            worksheet["S%d"%row].number_format = "dd-mm-yyyy"
            
            worksheet["T%d"%row] = line_id.related_document_type or ""
            serie = ''
            numero = ''
            if line_id.related_series:
                if len(line_id.related_number.split('-'))==1:
                    serie = line_id.related_number.split('/')[0]
                    numero = line_id.related_number[len(serie)+1:]
                else:
                    serie = line_id.related_number.split('-')[0]
                    numero = line_id.related_number[len(serie)+1:]
            worksheet["U%d"%row] = line_id.related_series or ""
            worksheet["V%d"%row] = line_id.related_number or ""
            row += 1
        if row == 12:
            row+=1    
        worksheet["J%d"%row] = "=SUM(J%d:J%d)"%(12, row-1)
        worksheet["K%d"%row] = "=SUM(K%d:K%d)"%(12, row-1)
        worksheet["L%d"%row] = "=SUM(L%d:L%d)"%(12, row-1)
        
        worksheet["M%d"%row] = "=SUM(M%d:M%d)"%(12, row-1)
        worksheet["N%d"%row] = "=SUM(N%d:N%d)"%(12, row-1)
        
        worksheet["O%d"%row] = "=SUM(O%d:O%d)"%(12, row-1)
        worksheet["P%d"%row] = "=SUM(P%d:P%d)"%(12, row-1)
        
        worksheet["Q%d"%row] = "=SUM(Q%d:Q%d)"%(12, row-1)
        
        
            